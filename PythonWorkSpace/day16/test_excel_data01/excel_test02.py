"""
--------------------------------
@Time    :  2022/2/25 - 14:56
@Author  :  Lee
@File    :  excel_test02
--------------------------------
"""

"""
python操作excel
"""

import openpyxl  # 用来操作excel

workbook = openpyxl.load_workbook(r'./cases.xlsx')  # 加载excel文件并且看作一个对象
sheet = workbook["user_login"]

# 获取cell中的值(excel文件读取的数据除了int/float,其他全是str字符串类型)
res = sheet.cell(row=2, column=2).value
print(res)  # 正常流程
res = sheet.cell(row=2, column=5).value
print(res)  # {"username": "xiaohua", "password": "a123456"}
print(type(res))  # <class 'str'>
res = sheet.cell(row=2, column=1).value
print(res)  # 1
print(type(res))  # <class 'int'>
# 获取所有的cell
print(list(sheet.rows))  # 把每一行有值的数据装在元组里返回

# 写数据
sheet.cell(row=2, column=7, value='tester24班的第一次excel读写!')
workbook.save(r'cases.xlsx')
