"""
--------------------------------
@Time    :  2022/2/25 - 16:21
@Author  :  Lee
@File    :  excel_utils03
--------------------------------
"""
import openpyxl

"""
将测试数据封装成对象
"""


class CaseData:
    def __init__(self, dict_case):
        for i in dict_case.items():
            setattr(self, i[0], i[1])


class ReadExcel:
    @classmethod
    def read_data_all(cls, file_name, sheet_name):  # 读取所有内容
        workbook = openpyxl.load_workbook(file_name)  # 读取工作簿
        sheet = workbook[sheet_name]  # 获取操作指定的sheet页
        rows = list(sheet.rows)  # 获取所有的单元格
        all_case = []

        # 读取表头数据
        titles = []
        for cell in rows[0]:  # rows[0]代表第一行数据
            titles.append(cell.value)  # 将第一行单元格的值添加到titles中
        # print(titles)  # ['case_id','case_title','interface','url'....]

        # 遍历其他行数据,和表头打包转换成字典,反射到对象的实例属性并把对象存到all_case列表里
        for row in rows[1:]:  # row代表了除了表头的每一行数据
            data = []
            for v in row:
                data.append(v.value)
            # print(data)  # [1,'正常流程','user_login'.....]
            case_data = dict(zip(titles, data))  # 将数据进行打包 {"case_id": 1 , "case_title": '正常流程'.....}
            cd = CaseData(case_data)  # 调用反射类,将字典转换为实例对象,字键的键就是表头,就是对象的实例属性
            all_case.append(cd)

        return all_case

    @classmethod
    def read_data_pl(cls, file_name, sheet_name, begin_row, end_row):  # 读取指定行数据
        workbook = openpyxl.load_workbook(file_name)  # 读取工作簿
        sheet = workbook[sheet_name]  # 获取操作指定的sheet页
        rows = list(sheet.rows)  # 获取所有的单元格
        all_case = []

        # 读取表头数据
        titles = []
        for cell in rows[0]:  # rows[0]代表第一行数据
            titles.append(cell.value)  # 将第一行单元格的值添加到titles中

        for row in rows[begin_row:end_row + 1]:  # row代表了除了表头的每一行数据
            data = []
            for v in row:
                data.append(v.value)
            case_data = dict(zip(titles, data))  # 将数据进行打包 {"case_id": 1 , "case_title": '正常流程'.....}
            cd = CaseData(case_data)  # 调用反射类,将字典转换为实例对象,字键的键就是表头,就是对象的实例属性
            all_case.append(cd)

        return all_case

    @classmethod
    def write_data(cls, file_name, sheet_name, row, column, value):  # 写数据
        """
        :param file_name: 需要被写入的文件
        :param sheet_name: 需要被写入的sheet页
        :param row: 需要被写入的行
        :param column: 需要被写入的列
        :param value: 需要写入的内容
        """
        wb = openpyxl.load_workbook(file_name)
        sh = wb[sheet_name]
        sh.cell(row=row + 1, column=column, value=value)
        wb.save(file_name)


if __name__ == '__main__':
    cases = ReadExcel.read_data_all(r'cases.xlsx', 'user_login')
    for case in cases:
        print(case.case_id, case.case_title, case.case_data)

    print('===' * 50)

    data = ReadExcel.read_data_pl(r'./cases.xlsx', 'user_login', 2, 2)
    print(data[0].case_id, data[0].case_title, data[0].url)

    ReadExcel.write_data(r'./cases.xlsx', 'user_login', 2, 7, '成功')
